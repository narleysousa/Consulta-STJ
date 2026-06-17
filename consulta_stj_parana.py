#!/usr/bin/env python3
"""
Consulta processos do STJ com origem no Paraná (TJPR) via API pública do Datajud.

Filtra processos com trânsito em julgado (TPU 848) e/ou baixa definitiva (TPU 22).
Suporta busca por período, por número de processo e busca em lote.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Callable, Iterable

API_URL = "https://api-publica.datajud.cnj.jus.br/api_publica_stj/_search"
API_KEY = "cDZHYzlZa0JadVREZDJCendQbXY6SkJlTzNjLV9TRENyQk1RdnFKZGRQdw=="

TRANSITO_JULGADO = 848
BAIXA_DEFINITIVA = 22

# Segmento 8 (Justiça Estadual) + tribunal 16 (Paraná/TJPR) no número CNJ.
TJPR_WILDCARD = "?????????????816????"

PAGE_SIZE = 1000
REQUEST_DELAY_SEC = 0.4
ANOS_PADRAO = range(2008, datetime.now().year + 1)

TIPOS_DATA = ("movimentacao", "ajuizamento", "atualizacao")

NOMES_MOVIMENTOS = {
    848: "Trânsito em julgado",
    22: "Baixa definitiva",
    123: "Remessa",
    132: "Recebimento",
    26: "Distribuição",
    51: "Conclusão",
    85: "Petição",
    92: "Publicação",
    54: "Acórdão",
    12: "Sentença",
    36: "Redistribuição",
}


@dataclass
class FiltroDatas:
    data_inicio: date | None = None
    data_fim: date | None = None
    tipo: str = "movimentacao"

    def ativo(self) -> bool:
        return self.data_inicio is not None and self.data_fim is not None

    def validar(self) -> None:
        if not self.ativo():
            return
        if self.data_inicio > self.data_fim:
            raise ValueError("A data inicial não pode ser maior que a data final.")
        if self.tipo not in TIPOS_DATA:
            raise ValueError(f"Tipo de data inválido: {self.tipo}")


@dataclass
class MovimentoResumo:
    codigo: int
    nome: str
    data_hora: str
    orgao: str = ""

    def data_formatada(self) -> str:
        return formatar_data_br(self.data_hora)


@dataclass
class ProcessoResultado:
    numero_processo: str
    numero_formatado: str
    classe: str
    assuntos: list[str]
    data_ajuizamento: str
    data_ultima_atualizacao: str
    tem_transito_julgado: bool
    data_transito_julgado: str
    tem_baixa_definitiva: bool
    data_baixa_definitiva: str
    timeline: list[MovimentoResumo] = field(default_factory=list)

    def as_dict(self) -> dict[str, str]:
        return {
            "numero_processo": self.numero_processo,
            "numero_formatado": self.numero_formatado,
            "classe": self.classe,
            "assuntos": " | ".join(self.assuntos),
            "data_ajuizamento": self.data_ajuizamento,
            "data_ultima_atualizacao": self.data_ultima_atualizacao,
            "tem_transito_julgado": "Sim" if self.tem_transito_julgado else "Não",
            "data_transito_julgado": self.data_transito_julgado,
            "tem_baixa_definitiva": "Sim" if self.tem_baixa_definitiva else "Não",
            "data_baixa_definitiva": self.data_baixa_definitiva,
        }


def corrigir_texto(texto: str) -> str:
    """Corrige mojibake comum da API (ex.: PRESIDÃ NCIA → PRESIDÊNCIA)."""
    if not texto:
        return ""
    texto = texto.strip()
    if any(m in texto for m in ("Ã", "Â", "\ufffd")):
        for encoding in ("latin-1", "cp1252"):
            try:
                texto = texto.encode(encoding).decode("utf-8")
                break
            except (UnicodeDecodeError, UnicodeEncodeError):
                continue
    for errado, certo in (
        ("PRESIDÃ NCIA", "PRESIDÊNCIA"),
        ("PRESIDÃNCIA", "PRESIDÊNCIA"),
        ("PRESIDÃŠNCIA", "PRESIDÊNCIA"),
    ):
        texto = texto.replace(errado, certo)
    return texto


def formatar_cnj(numero: str) -> str:
    n = numero.zfill(20)
    return f"{n[:7]}-{n[7:9]}.{n[9:13]}.{n[13]}.{n[14:16]}.{n[16:]}"


def formatar_data_br(valor: str) -> str:
    if not valor:
        return ""
    valor = valor.strip()
    if "T" in valor:
        try:
            return datetime.fromisoformat(valor.replace("Z", "+00:00")).strftime(
                "%d/%m/%Y %H:%M"
            )
        except ValueError:
            return valor
    if len(valor) >= 8 and valor[:8].isdigit():
        return f"{valor[6:8]}/{valor[4:6]}/{valor[0:4]}"
    return valor


def parse_para_data(valor: str) -> date | None:
    if not valor:
        return None
    valor = valor.strip()
    if "T" in valor:
        try:
            return datetime.fromisoformat(valor.replace("Z", "+00:00")).date()
        except ValueError:
            return None
    if len(valor) >= 8 and valor[:8].isdigit():
        try:
            return date(int(valor[0:4]), int(valor[4:6]), int(valor[6:8]))
        except ValueError:
            return None
    return None


def formatar_data_api_ajuizamento(d: date, fim_do_dia: bool = False) -> str:
    return f"{d.strftime('%Y%m%d')}{'235959' if fim_do_dia else '000000'}"


def formatar_data_api_iso(d: date, fim_do_dia: bool = False) -> str:
    return f"{d.isoformat()}{'T23:59:59' if fim_do_dia else 'T00:00:00'}"


def fatias_anuais(data_inicio: date, data_fim: date) -> list[tuple[date, date]]:
    fatias: list[tuple[date, date]] = []
    for ano in range(data_inicio.year, data_fim.year + 1):
        inicio = max(data_inicio, date(ano, 1, 1))
        fim = min(data_fim, date(ano, 12, 31))
        fatias.append((inicio, fim))
    return fatias


def api_buscar(body: dict, tentativas: int = 4) -> dict:
    payload = json.dumps(body).encode("utf-8")
    headers = {
        "Authorization": f"APIKey {API_KEY}",
        "Content-Type": "application/json",
    }
    for tentativa in range(1, tentativas + 1):
        req = urllib.request.Request(API_URL, data=payload, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                return json.load(resp)
        except urllib.error.HTTPError as exc:
            if tentativa == tentativas:
                raise RuntimeError(f"Erro HTTP {exc.code} na API Datajud") from exc
        except urllib.error.URLError as exc:
            if tentativa == tentativas:
                raise RuntimeError("Falha de conexão com a API Datajud") from exc
        time.sleep(REQUEST_DELAY_SEC * tentativa)
    raise RuntimeError("Não foi possível consultar a API Datajud")


def montar_query_base(
    periodo_inicio: date | None = None,
    periodo_fim: date | None = None,
    tipo_data: str = "movimentacao",
    modo: str = "qualquer",
) -> list[dict]:
    filtros: list[dict] = [
        {"wildcard": {"numeroProcesso": TJPR_WILDCARD}},
    ]
    if modo == "ambos":
        filtros.extend([
            {"term": {"movimentos.codigo": TRANSITO_JULGADO}},
            {"term": {"movimentos.codigo": BAIXA_DEFINITIVA}},
        ])
    elif modo == "transito":
        filtros.append({"term": {"movimentos.codigo": TRANSITO_JULGADO}})
    elif modo == "baixa":
        filtros.append({"term": {"movimentos.codigo": BAIXA_DEFINITIVA}})
    else:
        filtros.append({"terms": {"movimentos.codigo": [TRANSITO_JULGADO, BAIXA_DEFINITIVA]}})

    if periodo_inicio and periodo_fim:
        if tipo_data == "ajuizamento":
            filtros.append({"range": {"dataAjuizamento": {
                "gte": formatar_data_api_ajuizamento(periodo_inicio),
                "lte": formatar_data_api_ajuizamento(periodo_fim, fim_do_dia=True),
            }}})
        elif tipo_data == "atualizacao":
            filtros.append({"range": {"dataHoraUltimaAtualizacao": {
                "gte": formatar_data_api_iso(periodo_inicio),
                "lte": formatar_data_api_iso(periodo_fim, fim_do_dia=True),
            }}})
        else:
            # Movimentação: filtra server-side pela data dos movimentos.
            # (movimentos é object array → casa quando QUALQUER movimento cai no período)
            filtros.append({"range": {"movimentos.dataHora": {
                "gte": formatar_data_api_iso(periodo_inicio),
                "lte": formatar_data_api_iso(periodo_fim, fim_do_dia=True),
            }}})
    return filtros


def extrair_movimentos(movimentos: Iterable[dict], codigo: int) -> list[MovimentoResumo]:
    encontrados: list[MovimentoResumo] = []
    for m in movimentos:
        if not isinstance(m, dict):
            continue
        if m.get("codigo") == codigo:
            orgao = m.get("orgaoJulgador", {}) or {}
            if not isinstance(orgao, dict):
                orgao = {}
            encontrados.append(MovimentoResumo(
                codigo=codigo,
                nome=str(m.get("nome", "")),
                data_hora=str(m.get("dataHora", "")),
                orgao=corrigir_texto(str(orgao.get("nome", ""))),
            ))
    return encontrados


def extrair_timeline(movimentos: list[dict]) -> list[MovimentoResumo]:
    timeline: list[MovimentoResumo] = []
    for m in movimentos:
        if not isinstance(m, dict):
            continue
        try:
            codigo = int(m.get("codigo", 0))
        except (TypeError, ValueError):
            continue
        if codigo in NOMES_MOVIMENTOS:
            orgao = m.get("orgaoJulgador", {}) or {}
            if not isinstance(orgao, dict):
                orgao = {}
            timeline.append(MovimentoResumo(
                codigo=codigo,
                nome=str(m.get("nome", NOMES_MOVIMENTOS.get(codigo, ""))),
                data_hora=str(m.get("dataHora", "")),
                orgao=corrigir_texto(str(orgao.get("nome", ""))),
            ))
    timeline.sort(key=lambda x: x.data_hora)
    return timeline


def movimentos_no_periodo(
    movimentos: list[MovimentoResumo], data_inicio: date, data_fim: date
) -> list[MovimentoResumo]:
    return [
        m for m in movimentos
        if (d := parse_para_data(m.data_hora)) and data_inicio <= d <= data_fim
    ]


def data_no_periodo(valor: str, data_inicio: date, data_fim: date) -> bool:
    d = parse_para_data(valor)
    return bool(d and data_inicio <= d <= data_fim)


def atende_criterio(tem_transito: bool, tem_baixa: bool, modo: str) -> bool:
    if modo == "ambos":
        return tem_transito and tem_baixa
    if modo == "transito":
        return tem_transito
    if modo == "baixa":
        return tem_baixa
    if modo == "qualquer":
        return tem_transito or tem_baixa
    raise ValueError(f"Modo inválido: {modo}")


def extrair_assuntos(assuntos_raw) -> list[str]:
    """Extrai nomes de assuntos, tolerando dict único, lista ou aninhamento."""
    nomes: list[str] = []

    def coletar(item) -> None:
        if isinstance(item, dict):
            nome = item.get("nome")
            if nome:
                nomes.append(corrigir_texto(str(nome)))
        elif isinstance(item, list):
            for sub in item:
                coletar(sub)
        elif isinstance(item, str) and item.strip():
            nomes.append(item.strip())

    coletar(assuntos_raw)
    return nomes


def processar_hit(
    fonte: dict,
    modo: str,
    filtro_datas: FiltroDatas | None = None,
) -> ProcessoResultado | None:
    movimentos_raw = fonte.get("movimentos", [])
    transitos = extrair_movimentos(movimentos_raw, TRANSITO_JULGADO)
    baixas = extrair_movimentos(movimentos_raw, BAIXA_DEFINITIVA)

    # Presença geral dos movimentos (independente de data)
    tem_transito = bool(transitos)
    tem_baixa = bool(baixas)
    if not atende_criterio(tem_transito, tem_baixa, modo):
        return None

    if filtro_datas and filtro_datas.ativo():
        if filtro_datas.tipo == "movimentacao":
            # Exige que ao menos um dos movimentos relevantes ocorra no período.
            transitos_periodo = movimentos_no_periodo(transitos, filtro_datas.data_inicio, filtro_datas.data_fim)
            baixas_periodo = movimentos_no_periodo(baixas, filtro_datas.data_inicio, filtro_datas.data_fim)
            if not transitos_periodo and not baixas_periodo:
                return None
        elif filtro_datas.tipo == "ajuizamento":
            if not data_no_periodo(str(fonte.get("dataAjuizamento", "")), filtro_datas.data_inicio, filtro_datas.data_fim):
                return None
        elif filtro_datas.tipo == "atualizacao":
            if not data_no_periodo(str(fonte.get("dataHoraUltimaAtualizacao", "")), filtro_datas.data_inicio, filtro_datas.data_fim):
                return None

    numero = str(fonte.get("numeroProcesso", ""))
    ultimo_transito = max(transitos, key=lambda m: m.data_hora) if transitos else None
    ultima_baixa = max(baixas, key=lambda m: m.data_hora) if baixas else None

    assuntos = extrair_assuntos(fonte.get("assuntos"))

    classe_raw = fonte.get("classe")
    if isinstance(classe_raw, list):
        classe_raw = classe_raw[0] if classe_raw else {}
    classe_nome = corrigir_texto(str((classe_raw or {}).get("nome", ""))) if isinstance(classe_raw, dict) else ""

    return ProcessoResultado(
        numero_processo=numero,
        numero_formatado=formatar_cnj(numero),
        classe=classe_nome,
        assuntos=assuntos,
        data_ajuizamento=formatar_data_br(str(fonte.get("dataAjuizamento", ""))),
        data_ultima_atualizacao=formatar_data_br(str(fonte.get("dataHoraUltimaAtualizacao", ""))),
        tem_transito_julgado=tem_transito,
        data_transito_julgado=formatar_data_br(ultimo_transito.data_hora if ultimo_transito else ""),
        tem_baixa_definitiva=tem_baixa,
        data_baixa_definitiva=formatar_data_br(ultima_baixa.data_hora if ultima_baixa else ""),
        timeline=extrair_timeline(movimentos_raw),
    )


def buscar_por_periodo(
    periodo_inicio: date | None,
    periodo_fim: date | None,
    modo: str,
    filtro_datas: FiltroDatas | None = None,
    limite_restante: int | None = None,
) -> list[ProcessoResultado]:
    resultados: list[ProcessoResultado] = []
    offset = 0
    tipo_data = filtro_datas.tipo if filtro_datas and filtro_datas.ativo() else "movimentacao"

    while True:
        if limite_restante is not None and len(resultados) >= limite_restante:
            break

        tam = PAGE_SIZE if limite_restante is None else min(PAGE_SIZE, limite_restante - len(resultados))

        body = {
            "size": tam,
            "from": offset,
            "query": {"bool": {"must": montar_query_base(periodo_inicio, periodo_fim, tipo_data, modo)}},
            "_source": ["numeroProcesso", "classe", "assuntos", "dataAjuizamento", "dataHoraUltimaAtualizacao", "movimentos"],
        }
        dados = api_buscar(body)
        hits = dados.get("hits", {}).get("hits", [])
        if not hits:
            break

        for hit in hits:
            p = processar_hit(hit.get("_source", {}), modo, filtro_datas=filtro_datas)
            if p:
                resultados.append(p)
                if limite_restante is not None and len(resultados) >= limite_restante:
                    return resultados

        total = dados.get("hits", {}).get("total", {}).get("value", 0)
        offset += tam
        time.sleep(REQUEST_DELAY_SEC)
        if offset >= total or offset >= 10_000:
            break

    return resultados


def buscar_por_numero(numeros: list[str], modo: str = "qualquer") -> list[ProcessoResultado]:
    """Busca processos específicos por número CNJ (sem formatação)."""
    resultados: list[ProcessoResultado] = []
    limpos = [n.replace("-", "").replace(".", "") for n in numeros if n.strip()]

    for numero in limpos:
        body = {
            "size": 1,
            "query": {"term": {"numeroProcesso": numero}},
            "_source": ["numeroProcesso", "classe", "assuntos", "dataAjuizamento", "dataHoraUltimaAtualizacao", "movimentos"],
        }
        try:
            dados = api_buscar(body)
            hits = dados.get("hits", {}).get("hits", [])
            if hits:
                p = processar_hit(hits[0].get("_source", {}), modo)
                if p:
                    resultados.append(p)
        except RuntimeError:
            pass
        time.sleep(REQUEST_DELAY_SEC)

    return resultados


def buscar_processos(
    modo: str,
    limite: int | None,
    filtro_datas: FiltroDatas | None = None,
    anos: Iterable[int] | None = None,
    on_progress: Callable[[str, float], None] | None = None,
) -> list[ProcessoResultado]:
    if filtro_datas:
        filtro_datas.validar()

    vistos: set[str] = set()
    consolidado: list[ProcessoResultado] = []

    if filtro_datas and filtro_datas.ativo():
        periodos = fatias_anuais(filtro_datas.data_inicio, filtro_datas.data_fim)
    elif anos:
        periodos = [(date(a, 1, 1), date(a, 12, 31)) for a in anos]
    else:
        periodos = [(date(a, 1, 1), date(a, 12, 31)) for a in ANOS_PADRAO]

    total_periodos = len(periodos)
    for idx, (pi, pf) in enumerate(periodos):
        pct = idx / total_periodos
        msg = f"Buscando {pi.strftime('%d/%m/%Y')} → {pf.strftime('%d/%m/%Y')} (período {idx+1}/{total_periodos})"
        if on_progress:
            on_progress(msg, pct)
        else:
            print(msg, file=sys.stderr)

        restante = None if limite is None else max(limite - len(consolidado), 0)
        if restante == 0:
            break

        fd_periodo = filtro_datas
        if filtro_datas and filtro_datas.ativo():
            fd_periodo = FiltroDatas(
                data_inicio=max(filtro_datas.data_inicio, pi),
                data_fim=min(filtro_datas.data_fim, pf),
                tipo=filtro_datas.tipo,
            )

        for p in buscar_por_periodo(pi, pf, modo, filtro_datas=fd_periodo, limite_restante=restante):
            if p.numero_processo not in vistos:
                vistos.add(p.numero_processo)
                consolidado.append(p)
                if limite is not None and len(consolidado) >= limite:
                    if on_progress:
                        on_progress("Limite atingido.", 1.0)
                    return consolidado

    if on_progress:
        on_progress(f"Concluído — {len(consolidado)} processos encontrados.", 1.0)
    consolidado.sort(key=lambda x: x.numero_processo)
    return consolidado


def processos_para_csv(processos: list[ProcessoResultado]) -> str:
    campos = list(ProcessoResultado("", "", "", [], "", "", False, "", False, "").as_dict().keys())
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=campos)
    writer.writeheader()
    for p in processos:
        writer.writerow(p.as_dict())
    return "\ufeff" + buf.getvalue()


def processos_para_excel(processos: list[ProcessoResultado]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Instale openpyxl: pip install openpyxl") from exc

    campos_map = {
        "numero_formatado": "Número do Processo",
        "classe": "Classe",
        "assuntos": "Assuntos",
        "data_ajuizamento": "Ajuizamento",
        "data_transito_julgado": "Trânsito em Julgado",
        "data_baixa_definitiva": "Baixa Definitiva",
        "data_ultima_atualizacao": "Última Atualização",
        "tem_transito_julgado": "Tem Trânsito",
        "tem_baixa_definitiva": "Tem Baixa",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "STJ Paraná"

    header_fill = PatternFill("solid", fgColor="0F3D5C")
    header_font = Font(color="FFFFFF", bold=True)
    for col, nome in enumerate(campos_map.values(), 1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for p in processos:
        d = p.as_dict()
        ws.append([d.get(k, "") for k in campos_map])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_csv(caminho: Path, processos: list[ProcessoResultado]) -> None:
    caminho.write_text(processos_para_csv(processos), encoding="utf-8-sig")


def exportar_excel(caminho: Path, processos: list[ProcessoResultado]) -> None:
    caminho.write_bytes(processos_para_excel(processos))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Lista processos STJ com origem no Paraná (trânsito em julgado / baixa definitiva)."
    )
    parser.add_argument("--modo", choices=["ambos", "transito", "baixa", "qualquer"], default="ambos")
    parser.add_argument("--data-inicio", help="Data inicial AAAA-MM-DD")
    parser.add_argument("--data-fim", help="Data final AAAA-MM-DD")
    parser.add_argument("--tipo-data", choices=list(TIPOS_DATA), default="movimentacao")
    parser.add_argument("--ano", type=int, action="append", help="Filtrar por ano de ajuizamento")
    parser.add_argument("--limite", type=int)
    parser.add_argument("--saida", type=Path, default=Path("stj_parana_resultados.csv"))
    parser.add_argument("--excel", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    filtro_datas: FiltroDatas | None = None

    if args.data_inicio and args.data_fim:
        filtro_datas = FiltroDatas(
            data_inicio=datetime.strptime(args.data_inicio, "%Y-%m-%d").date(),
            data_fim=datetime.strptime(args.data_fim, "%Y-%m-%d").date(),
            tipo=args.tipo_data,
        )
        filtro_datas.validar()
    elif args.ano:
        anos_ord = sorted(args.ano)
        filtro_datas = FiltroDatas(
            data_inicio=date(min(anos_ord), 1, 1),
            data_fim=date(max(anos_ord), 12, 31),
            tipo="ajuizamento",
        )

    print("Consultando API pública do Datajud (STJ / Paraná)...", file=sys.stderr)
    processos = buscar_processos(modo=args.modo, limite=args.limite, filtro_datas=filtro_datas)

    if not processos:
        print("Nenhum processo encontrado.", file=sys.stderr)
        return 1

    exportar_csv(args.saida, processos)
    print(f"CSV: {args.saida.resolve()} ({len(processos)} processos)", file=sys.stderr)
    if args.excel:
        exportar_excel(args.excel, processos)
        print(f"Excel: {args.excel.resolve()}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
